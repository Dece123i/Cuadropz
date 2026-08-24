import sqlite3
import datetime
import os
import re
import json
import openpyxl

DB_PATH = "cuadropz.db"

def db_connect():
    return sqlite3.connect(DB_PATH)

def db_init():
    conn = db_connect()
    cursor = conn.cursor()
    
    # Create tasks table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_name TEXT NOT NULL,
        date TEXT NOT NULL,
        description TEXT NOT NULL,
        time_info TEXT,
        priority TEXT DEFAULT 'Normal',
        completed INTEGER DEFAULT 0,
        order_num INTEGER NOT NULL,
        carried_over_from TEXT
    )
    """)
    
    # Create reports table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_name TEXT NOT NULL,
        date TEXT NOT NULL,
        resolved_tasks TEXT,
        unresolved_tasks TEXT,
        alternatives_of_solution TEXT,
        execution_status TEXT
    )
    """)
    
    # Create user_state table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS user_state (
        user_name TEXT PRIMARY KEY,
        last_finalized_date TEXT
    )
    """)
    
    # Create users table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        name TEXT PRIMARY KEY,
        is_admin INTEGER DEFAULT 0
    )
    """)
    
    # Migration for existing DBs
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER DEFAULT 0")
    except sqlite3.OperationalError:
        pass
        
    try:
        cursor.execute("ALTER TABLE reports ADD COLUMN execution_status TEXT")
    except sqlite3.OperationalError:
        pass
        
    # Insert default users if empty
    cursor.execute("SELECT COUNT(*) FROM users")
    if cursor.fetchone()[0] == 0:
        cursor.executemany("INSERT INTO users (name, is_admin) VALUES (?, ?)", [("MARY CRUZ", 0), ("CPC.SHEYLA", 0), ("CPC.HECTOR", 0)])
        
    conn.commit()
    conn.close()

def get_users():
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM users ORDER BY name ASC")
    rows = cursor.fetchall()
    conn.close()
    return [r[0] for r in rows]

def add_user(name):
    conn = db_connect()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO users (name) VALUES (?)", (name.strip(),))
        conn.commit()
        success = True
    except sqlite3.IntegrityError:
        success = False
    conn.close()
    return success

def delete_user(name):
    conn = db_connect()
    cursor = conn.cursor()
    try:
        cursor.execute("BEGIN TRANSACTION")
        cursor.execute("DELETE FROM tasks WHERE user_name = ?", (name,))
        cursor.execute("DELETE FROM reports WHERE user_name = ?", (name,))
        cursor.execute("DELETE FROM user_state WHERE user_name = ?", (name,))
        cursor.execute("DELETE FROM users WHERE name = ?", (name,))
        cursor.execute("COMMIT")
        success = True
    except Exception as e:
        cursor.execute("ROLLBACK")
        print("Delete user error:", e)
        success = False
    conn.close()
    return success

def parse_task_text(text):
    """
    Parses a cell text from the Excel sheet into (description, time_info).
    E.g. "1.PROGRAMACION DE QOORI - 8.00AM A 9.00 AM" 
    -> ("PROGRAMACION DE QOORI", "8.00AM A 9.00 AM")
    """
    if not text:
        return None, None
    text = str(text).strip()
    if text.upper() == "NONE" or text == "":
        return None, None
        
    # Strip leading index e.g. "1.", "1. "
    text = re.sub(r'^\d+[\.\s]+', '', text).strip()
    
    # Find the last hyphen in the string
    last_hyphen_idx = text.rfind(' - ')
    if last_hyphen_idx != -1:
        potential_time = text[last_hyphen_idx + 3:].strip()
        # Verify if it looks like a time: contains digits and either AM/PM or 'A' or 'HASTA' or '.' or ':'
        if re.search(r'\d', potential_time) and any(x in potential_time.upper() for x in ['AM', 'PM', 'A ', 'HASTA', '.', ':']):
            description = text[:last_hyphen_idx].strip()
            return description, potential_time

    # General time pattern at the end (if no hyphen was found or didn't look like a time)
    # The time pattern typically consists of: <time1> <sep> <time2> where <sep> is HASTA or A
    time_pattern = r'(\d+[\d\.:\s]*(?:AM|PM|am|pm|am|pm)?\s+(?:HASTA|A|a)\s+\d+[\d\.:\s]*(?:AM|PM|am|pm|am|pm)?)$'
    match_time = re.search(time_pattern, text, re.IGNORECASE)
    if match_time:
        time_info = match_time.group(1).strip()
        if re.search(r'\d', time_info):
            description = text[:match_time.start()].strip()
            if description.endswith('-'):
                description = description[:-1].strip()
            return description, time_info
            
    return text, None

def db_import_excel(excel_path):
    """
    Parses the "CUADRO DE PRODUCCION" sheet and imports all tasks into SQLite tasks table.
    Only runs if the tasks table is empty to avoid double imports.
    """
    conn = db_connect()
    cursor = conn.cursor()
    
    # Check if table already has tasks
    cursor.execute("SELECT COUNT(*) FROM tasks")
    if cursor.fetchone()[0] > 0:
        conn.close()
        return False # Already imported or has data
        
    if not os.path.exists(excel_path):
        conn.close()
        return False
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "CUADRO DE PRODUCCION" not in wb.sheetnames:
            conn.close()
            return False
            
        sheet = wb["CUADRO DE PRODUCCION"]
        max_row = sheet.max_row
        
        # 1. Find all weekly headers (where Column 2 contains "NOMBRES")
        header_rows = []
        for r in range(1, max_row + 1):
            val = sheet.cell(r, 2).value
            if val and "NOMBRES" in str(val).upper():
                header_rows.append(r)
                
        # 2. Parse each week
        for idx, h in enumerate(header_rows):
            # Determine the end row of this week's block
            h_next = header_rows[idx + 1] if idx + 1 < len(header_rows) else max_row + 1
            
            # Parse the Monday date from Column 3
            monday_cell = sheet.cell(h, 3).value
            if not monday_cell:
                continue
            
            # Extract date string like "20/04" or similar
            match = re.search(r'(\d{2})/(\d{2})', str(monday_cell))
            if not match:
                continue
                
            day = int(match.group(1))
            month = int(match.group(2))
            # Assume year 2026 as found in files
            try:
                monday_date = datetime.date(2026, month, day)
            except ValueError:
                # Fallback if invalid date
                continue
                
            # Maps column index (3 to 8) to date strings (YYYY-MM-DD)
            col_dates = {}
            for col_idx in range(3, 9):
                offset = col_idx - 3
                target_date = monday_date + datetime.timedelta(days=offset)
                col_dates[col_idx] = target_date.strftime("%Y-%m-%d")
                
            # Scan Column 2 for user block rows within [h + 1, h_next - 1]
            row_sheyla = None
            row_hector = None
            
            for r in range(h + 1, h_next):
                val2 = sheet.cell(r, 2).value
                if val2:
                    val2_str = str(val2).upper()
                    if "SHEYLA" in val2_str:
                        row_sheyla = r
                    elif "HECTOR" in val2_str:
                        row_hector = r
            
            # If we couldn't find them, use standard offsets relative to h
            if not row_sheyla:
                row_sheyla = h + 10
            if not row_hector:
                row_hector = h + 15
                
            # Define row ranges for each user
            # MARY CRUZ: from h + 1 to row_sheyla - 1
            # CPC.SHEYLA: from row_sheyla to row_hector - 1
            # CPC.HECTOR: from row_hector onwards until tasks end (all columns empty or legend starts)
            
            user_ranges = [
                ("MARY CRUZ", h + 1, row_sheyla),
                ("CPC.SHEYLA", row_sheyla, row_hector),
                ("CPC.HECTOR", row_hector, h_next)
            ]
            
            for user_name, start_r, end_r in user_ranges:
                for r in range(start_r, end_r):
                    # Check if this row is empty or a legend row
                    col2_val = sheet.cell(r, 2).value
                    if col2_val and any(x in str(col2_val).upper() for x in ["LEYENDA", "ROJO", "ANARANJADO", "AMARILLO", "VERDE"]):
                        break
                        
                    # Check if all columns are empty
                    all_empty = True
                    for col_idx in range(3, 9):
                        c_val = sheet.cell(r, col_idx).value
                        if c_val and str(c_val).strip() != "" and str(c_val).upper() != "NONE":
                            all_empty = False
                            break
                    if all_empty:
                        # If this is HECTOR block and we hit an empty row, we stop
                        if user_name == "CPC.HECTOR":
                            break
                        continue
                        
                    # For each day (column index 3 to 8)
                    for col_idx in range(3, 9):
                        cell_val = sheet.cell(r, col_idx).value
                        if not cell_val:
                            continue
                        cell_val_str = str(cell_val).strip()
                        if cell_val_str == "" or cell_val_str.upper() == "NONE" or "FERIADO" in cell_val_str.upper():
                            continue
                            
                        # Parse task text
                        desc, time_info = parse_task_text(cell_val_str)
                        if not desc:
                            continue
                            
                        task_date = col_dates[col_idx]
                        
                        # Determine order_num for this day
                        cursor.execute("""
                            SELECT COALESCE(MAX(order_num), 0) + 1 
                            FROM tasks 
                            WHERE user_name = ? AND date = ?
                        """, (user_name, task_date))
                        order_num = cursor.fetchone()[0]
                        
                        # Check color of the cell to set priority and completed status
                        cell = sheet.cell(r, col_idx)
                        completed = 0
                        priority = 'Normal'
                        
                        # Get background color in hex
                        fg_color = None
                        if cell.fill and cell.fill.start_color:
                            rgb_val = cell.fill.start_color.rgb
                            if isinstance(rgb_val, str) and len(rgb_val) == 8:
                                fg_color = rgb_val
                            
                        # Map colors
                        # In openpyxl, a color is an 8-char hex (ARGB), e.g. "FFFF0000" for Red
                        # Sometimes it is 00000000 or similar
                        if fg_color:
                            color_hex = fg_color[2:].upper() # Strip alpha
                            
                            # Yellow (Amarillo) = completed
                            # Common Excel yellow hexes: FFFF00, FFE699, FFFFCC, FFF2CC, FFFF99
                            if color_hex in ["FFFF00", "FFE699", "FFF2CC", "FFFF99"]:
                                completed = 1
                            # Red (Rojo) = Alta
                            elif color_hex in ["FF0000", "FFC7CE", "F8CBAD", "FCE4D6"]:
                                priority = 'Alta'
                            # Orange (Anaranjado) = Media
                            elif color_hex in ["FFC000", "ED7D31", "FFF2CC"]:
                                priority = 'Media'
                            # Green (Verde) = Baja
                            elif color_hex in ["00B050", "385723", "C6EFCE", "E2EFDA"]:
                                priority = 'Baja'
                        
                        cursor.execute("""
                            INSERT INTO tasks (user_name, date, description, time_info, priority, completed, order_num)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (user_name, task_date, desc, time_info, priority, completed, order_num))
                        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print("Excel Import Error:", e)
        conn.close()
        return False

# TASK CRUD OPERATIONS
def get_tasks(user_name, date):
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, user_name, date, description, time_info, priority, completed, order_num, carried_over_from
        FROM tasks
        WHERE user_name = ? AND date = ?
        ORDER BY order_num ASC
    """, (user_name, date))
    rows = cursor.fetchall()
    conn.close()
    
    tasks = []
    for r in rows:
        tasks.append({
            'id': r[0],
            'user_name': r[1],
            'date': r[2],
            'description': r[3],
            'time_info': r[4],
            'priority': r[5],
            'completed': r[6],
            'order_num': r[7],
            'carried_over_from': r[8]
        })
    return tasks

def add_task(user_name, date, description, time_info=None, priority='Normal'):
    conn = db_connect()
    cursor = conn.cursor()
    
    # Get next order_num
    cursor.execute("""
        SELECT COALESCE(MAX(order_num), 0) + 1 
        FROM tasks 
        WHERE user_name = ? AND date = ?
    """, (user_name, date))
    order_num = cursor.fetchone()[0]
    
    cursor.execute("""
        INSERT INTO tasks (user_name, date, description, time_info, priority, completed, order_num)
        VALUES (?, ?, ?, ?, ?, 0, ?)
    """, (user_name, date, description, time_info, priority, order_num))
    
    conn.commit()
    conn.close()

def update_task_completion(task_id, completed):
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("UPDATE tasks SET completed = ? WHERE id = ?", (completed, task_id))
    conn.commit()
    conn.close()

def update_task_priority(task_id, priority):
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("UPDATE tasks SET priority = ? WHERE id = ?", (priority, task_id))
    conn.commit()
    conn.close()

def update_task_details(task_id, description, time_info):
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("UPDATE tasks SET description = ?, time_info = ? WHERE id = ?", (description, time_info, task_id))
    conn.commit()
    conn.close()

def delete_task(task_id):
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    conn.commit()
    conn.close()

# DATE HELPER AND CARRY OVER LOGIC
def get_next_working_day(date_str):
    """
    Returns the next working day (Lunes to Sábado).
    If it is Saturday, returns the next Monday.
    Otherwise, returns the next calendar day.
    """
    current_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    # Add 1 day
    next_day = current_date + datetime.timedelta(days=1)
    
    # If Sunday (6 in python weekday), add 1 more day to get Monday
    if next_day.weekday() == 6:
        next_day = next_day + datetime.timedelta(days=1)
        
    return next_day.strftime("%Y-%m-%d")

def finalize_day(user_name, date_str, resolved_list, unresolved_list, alternatives_list):
    """
    Saves the daily report, marks the day as finalized, and performs carry-over.
    """
    conn = db_connect()
    cursor = conn.cursor()
    
    # Ensure each unresolved task dictionary has an 'execution_status' key initialized to ''
    for task in unresolved_list:
        if isinstance(task, dict):
            task['execution_status'] = ""
            
    # Default execution status is empty for each unresolved task
    execution_status_list = [""] * len(unresolved_list)
    
    # 1. Save Report
    cursor.execute("""
        INSERT INTO reports (user_name, date, resolved_tasks, unresolved_tasks, alternatives_of_solution, execution_status)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (user_name, date_str, json.dumps(resolved_list), json.dumps(unresolved_list), json.dumps(alternatives_list), json.dumps(execution_status_list)))
    
    # 2. Update user last finalized date
    cursor.execute("""
        INSERT OR REPLACE INTO user_state (user_name, last_finalized_date)
        VALUES (?, ?)
    """, (user_name, date_str))
    
    # 3. Carry over unresolved tasks to next working day
    if len(unresolved_list) > 0:
        next_w_day = get_next_working_day(date_str)
        
        for task in unresolved_list:
            # Check for duplicates on the target day
            # task is a dict: {'description': ..., 'time_info': ..., 'priority': ...}
            desc = task['description']
            time_info = task.get('time_info')
            priority = task.get('priority', 'Normal')
            
            # Check if identical task already exists on target day for this user
            cursor.execute("""
                SELECT COUNT(*) FROM tasks 
                WHERE user_name = ? AND date = ? AND LOWER(TRIM(description)) = LOWER(TRIM(?))
            """, (user_name, next_w_day, desc))
            
            exists = cursor.fetchone()[0] > 0
            if not exists:
                # Insert task for next day
                cursor.execute("""
                    SELECT COALESCE(MAX(order_num), 0) + 1 
                    FROM tasks 
                    WHERE user_name = ? AND date = ?
                """, (user_name, next_w_day))
                next_order = cursor.fetchone()[0]
                
                cursor.execute("""
                    INSERT INTO tasks (user_name, date, description, time_info, priority, completed, order_num, carried_over_from)
                    VALUES (?, ?, ?, ?, ?, 0, ?, ?)
                """, (user_name, next_w_day, desc, time_info, priority, next_order, date_str))
                
    conn.commit()
    conn.close()

def get_last_finalized_date(user_name):
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("SELECT last_finalized_date FROM user_state WHERE user_name = ?", (user_name,))
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else None

def get_undone_days_before(user_name, date_str):
    """
    Finds dates strictly before `date_str` that have tasks for this user,
    but have not been finalized (i.e. date is greater than last_finalized_date, or no report exists).
    Returns a sorted list of these dates.
    """
    last_finalized = get_last_finalized_date(user_name)
    
    conn = db_connect()
    cursor = conn.cursor()
    
    # We select all distinct dates with tasks for this user that are < date_str
    if last_finalized:
        cursor.execute("""
            SELECT DISTINCT date FROM tasks
            WHERE user_name = ? AND date < ? AND date > ?
            ORDER BY date ASC
        """, (user_name, date_str, last_finalized))
    else:
        cursor.execute("""
            SELECT DISTINCT date FROM tasks
            WHERE user_name = ? AND date < ?
            ORDER BY date ASC
        """, (user_name, date_str))
        
    dates_with_tasks = [r[0] for r in cursor.fetchall()]
    
    # Filter out dates that already have a report just in case
    undone_dates = []
    for d in dates_with_tasks:
        cursor.execute("SELECT COUNT(*) FROM reports WHERE user_name = ? AND date = ?", (user_name, d))
        if cursor.fetchone()[0] == 0:
            undone_dates.append(d)
            
    conn.close()
    return undone_dates

def get_reports(user_name=None, start_date=None, end_date=None):
    conn = db_connect()
    cursor = conn.cursor()
    
    query = "SELECT id, user_name, date, resolved_tasks, unresolved_tasks, alternatives_of_solution, execution_status FROM reports WHERE 1=1"
    params = []
    
    if user_name:
        query += " AND user_name = ?"
        params.append(user_name)
    if start_date:
        query += " AND date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date <= ?"
        params.append(end_date)
        
    query += " ORDER BY date DESC, user_name ASC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    reports = []
    for r in rows:
        unresolved = json.loads(r[4]) if r[4] else []
        n = len(unresolved)
        status_list = []
        if len(r) > 6 and r[6]:
            try:
                status_list = json.loads(r[6])
            except Exception:
                status_list = [""] * n
        else:
            status_list = [""] * n
            
        # Align lengths of status_list and unresolved
        if len(status_list) < n:
            status_list.extend([""] * (n - len(status_list)))
        elif len(status_list) > n:
            status_list = status_list[:n]
            
        # Add execution_status key directly to each unresolved task dict for easy consumption
        for idx, task in enumerate(unresolved):
            if isinstance(task, dict):
                task['execution_status'] = status_list[idx]
                
        reports.append({
            'id': r[0],
            'user_name': r[1],
            'date': r[2],
            'resolved_tasks': json.loads(r[3]) if r[3] else [],
            'unresolved_tasks': unresolved,
            'alternatives_of_solution': json.loads(r[5]) if r[5] else [],
            'execution_status': status_list
        })
    return reports

def update_task_execution_status(report_id, task_index, status):
    """
    Updates the execution status (motivo) of a specific unresolved task in a report.
    """
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("SELECT unresolved_tasks, execution_status FROM reports WHERE id = ?", (report_id,))
    row = cursor.fetchone()
    if row:
        unresolved_tasks = json.loads(row[0]) if row[0] else []
        n = len(unresolved_tasks)
        if row[1]:
            try:
                status_list = json.loads(row[1])
            except Exception:
                status_list = [""] * n
        else:
            status_list = [""] * n
            
        # Adjust length of status_list if needed
        if len(status_list) < n:
            status_list.extend([""] * (n - len(status_list)))
        elif len(status_list) > n:
            status_list = status_list[:n]
            
        if 0 <= task_index < n:
            status_list[task_index] = status
            if isinstance(unresolved_tasks[task_index], dict):
                unresolved_tasks[task_index]['execution_status'] = status
            
        cursor.execute("""
            UPDATE reports 
            SET execution_status = ?, unresolved_tasks = ? 
            WHERE id = ?
        """, (json.dumps(status_list), json.dumps(unresolved_tasks), report_id))
        conn.commit()
    conn.close()

def get_days_with_pending_tasks(user_name):
    """
    Returns a dictionary of dates that have tasks for this user, mapped to their count of pending tasks.
    """
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT date, COUNT(*) 
        FROM tasks 
        WHERE user_name = ? AND completed = 0 
        GROUP BY date
    """, (user_name,))
    rows = cursor.fetchall()
    conn.close()
    return {r[0]: r[1] for r in rows}

def get_overdue_pending_tasks(user_name, date_str):
    """
    Returns a list of uncompleted tasks for dates prior to date_str.
    """
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, user_name, date, description, time_info, priority, completed, order_num, carried_over_from
        FROM tasks
        WHERE user_name = ? AND date < ? AND completed = 0
        ORDER BY date DESC, order_num ASC
    """, (user_name, date_str))
    rows = cursor.fetchall()
    conn.close()
    
    return [{
        'id': r[0],
        'user_name': r[1],
        'date': r[2],
        'description': r[3],
        'time_info': r[4],
        'priority': r[5],
        'completed': r[6],
        'order_num': r[7],
        'carried_over_from': r[8]
    } for r in rows]

def get_tasks_by_user_and_date(user_id, date):
    """
    Returns tasks for a given user and date.
    """
    if isinstance(date, (datetime.date, datetime.datetime)):
        date = date.strftime("%Y-%m-%d")
    return get_tasks(user_id, date)

def get_completed_tasks_count(user_id, date):
    """
    Returns the count of completed tasks for a given user and date.
    """
    if isinstance(date, (datetime.date, datetime.datetime)):
        date = date.strftime("%Y-%m-%d")
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT COUNT(*) FROM tasks
        WHERE user_name = ? AND date = ? AND completed = 1
    """, (user_id, date))
    count = cursor.fetchone()[0]
    conn.close()
    return count

def get_last_report(user_id):
    """
    Returns the last finalized report for a given user.
    """
    reports = get_reports(user_name=user_id)
    return reports[0] if reports else None

def save_last_export_info(date_str, count):
    """
    Saves metadata about the last Excel export to a local JSON file.
    """
    try:
        with open("last_export_info.json", "w") as f:
            json.dump({"date": date_str, "tasks_count": count}, f)
    except Exception as e:
        print("Error saving export info:", e)

def get_last_export_info():
    """
    Retrieves metadata about the last Excel export.
    """
    if os.path.exists("last_export_info.json"):
        try:
            with open("last_export_info.json", "r") as f:
                return json.load(f)
        except Exception as e:
            print("Error reading export info:", e)
    return None

def update_report_suggestions(report_id, suggestions_list):
    """
    Updates the alternatives_of_solution list in reports table for a given report ID.
    """
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE reports
        SET alternatives_of_solution = ?
        WHERE id = ?
    """, (json.dumps(suggestions_list), report_id))
    conn.commit()
    conn.close()

def get_all_users_with_admin_status():
    """
    Returns all users with their admin status.
    """
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("SELECT name, is_admin FROM users ORDER BY name ASC")
    rows = cursor.fetchall()
    conn.close()
    return [{"name": r[0], "is_admin": r[1]} for r in rows]

def toggle_admin_status(user_name):
    """
    Toggles the is_admin status of a user.
    """
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET is_admin = 1 - is_admin WHERE name = ?", (user_name,))
    conn.commit()
    conn.close()

def delete_user_by_admin(user_name):
    """
    Deletes a user from all tables.
    """
    return delete_user(user_name)

def add_user_with_role(name, is_admin):
    """
    Adds a new user with a specific admin role status.
    """
    conn = db_connect()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO users (name, is_admin) VALUES (?, ?)", (name.strip(), 1 if is_admin else 0))
        conn.commit()
        success = True
    except sqlite3.IntegrityError:
        success = False
    conn.close()
    return success


