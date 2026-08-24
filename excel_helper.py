import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from database import get_tasks, get_reports, get_users
from ai_helper import generate_execution_reason

def get_week_dates(monday_str):
    """
    Given a Monday date string (YYYY-MM-DD), 
    returns a list of 6 date strings (Lunes to Sábado).
    """
    monday = datetime.datetime.strptime(monday_str, "%Y-%m-%d").date()
    return [(monday + datetime.timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6)]

def export_weekly_tasks_to_excel(monday_str, output_path):
    """
    Generates a weekly Excel sheet in output_path containing tasks for all users,
    formatted similarly to the "CUADRO DE PRODUCCION" template.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create CUADRO DE PRODUCCION sheet
    sheet = wb.create_sheet(title="CUADRO DE PRODUCCION")
    
    # Enable grid lines visibility
    sheet.views.sheetView[0].showGridLines = True
    
    # Get week dates
    week_dates = get_week_dates(monday_str)
    
    # Formats the date to DD/MM
    def fmt_date(d_str):
        d = datetime.datetime.strptime(d_str, "%Y-%m-%d").date()
        return d.strftime("%d/%m")
        
    # Styles definition
    font_header = Font(name="Calibri", size=11, bold=True, color="000000")
    font_user = Font(name="Calibri", size=11, bold=True, color="000000")
    font_task = Font(name="Calibri", size=10, color="000000")
    font_legend = Font(name="Calibri", size=10, italic=True, color="595959")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border_thin_side = Side(style='thin', color='D9D9D9')
    border_task = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    
    border_header_side = Side(style='medium', color='7F7F7F')
    border_header = Border(left=border_thin_side, right=border_thin_side, top=border_header_side, bottom=border_header_side)
    
    # Priority fills (soft pastel colors)
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_completed = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow (Amarillo)
    fill_alta = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")      # Soft Red (Rojo)
    fill_media = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")     # Soft Orange (Anaranjado) - wait, let's use FDE9D9 for soft Orange
    fill_media = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    fill_baja = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")      # Soft Green (Verde)
    fill_normal = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # 1. Headers on Row 3
    # Row 1 and 2 are empty as in the original
    header_row = 3
    sheet.cell(header_row, 2, "NOMBRES").font = font_header
    sheet.cell(header_row, 2).alignment = align_center
    sheet.cell(header_row, 2).fill = fill_header
    sheet.cell(header_row, 2).border = border_header
    
    day_names = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]
    for i, date_str in enumerate(week_dates):
        col_idx = 3 + i
        col_header = f"{day_names[i]} {fmt_date(date_str)}"
        cell = sheet.cell(header_row, col_idx, col_header)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_header
        
    # 2. Add tasks for each user
    users = ["MARY CRUZ", "CPC.SHEYLA", "CPC.HECTOR"]
    current_row = 4
    
    for user_name in users:
        # Load tasks for each day
        user_tasks_by_day = {}
        max_tasks = 0
        for i, date_str in enumerate(week_dates):
            tasks = get_tasks(user_name, date_str)
            user_tasks_by_day[i] = tasks
            max_tasks = max(max_tasks, len(tasks))
            
        # Allocate rows for this user (at least 9 rows)
        allocated_rows = max(9, max_tasks)
        start_u_row = current_row
        end_u_row = current_row + allocated_rows - 1
        
        # Write user name and merge cells in Column B
        sheet.cell(start_u_row, 2, user_name).font = font_user
        sheet.cell(start_u_row, 2).alignment = align_center
        sheet.cell(start_u_row, 2).border = border_task
        if allocated_rows > 1:
            sheet.merge_cells(start_row=start_u_row, start_column=2, end_row=end_u_row, end_column=2)
            
            # Apply border to all merged cells in Column B
            for r in range(start_u_row, end_u_row + 1):
                sheet.cell(r, 2).border = border_task
                
        # Fill tasks in columns C to H
        for i, date_str in enumerate(week_dates):
            col_idx = 3 + i
            day_tasks = user_tasks_by_day[i]
            
            for task_idx in range(allocated_rows):
                r = start_u_row + task_idx
                cell = sheet.cell(r, col_idx)
                cell.border = border_task
                cell.alignment = align_left
                cell.font = font_task
                
                if task_idx < len(day_tasks):
                    task = day_tasks[task_idx]
                    # Format cell value: "1. Tarea - Hora" or "1. Tarea"
                    t_desc = task['description']
                    t_time = task['time_info']
                    
                    val_str = f"{task_idx + 1}. {t_desc}"
                    if t_time and str(t_time).strip() != "" and str(t_time).upper() != "NONE":
                        val_str += f" - {t_time}"
                        
                    cell.value = val_str
                    
                    # Apply colors based on completion or priority
                    if task['completed'] == 1:
                        cell.fill = fill_completed
                    else:
                        p = task['priority']
                        if p == 'Alta':
                            cell.fill = fill_alta
                        elif p == 'Media':
                            cell.fill = fill_media
                        elif p == 'Baja':
                            cell.fill = fill_baja
                        else:
                            cell.fill = fill_normal
                else:
                    # Empty cell
                    cell.value = ""
                    cell.fill = fill_normal
                    
        # Update current row for next user
        current_row = end_u_row + 1
        
    # 3. Add Legend Section
    current_row += 3  # Leave some blank rows
    
    legend_title_cell = sheet.cell(current_row, 2, "leyenda :")
    legend_title_cell.font = font_header
    
    current_row += 2
    
    # Red (Muy Urgente)
    c_red_lbl = sheet.cell(current_row, 2, "ROJO")
    c_red_lbl.font = font_legend
    c_red_lbl.fill = fill_alta
    c_red_lbl.alignment = align_center
    c_red_lbl.border = border_task
    sheet.cell(current_row, 4, "MUY URGENTE").font = font_legend
    
    current_row += 1
    
    # Orange (Urgente)
    c_orange_lbl = sheet.cell(current_row, 2, "ANARANJADO :")
    c_orange_lbl.font = font_legend
    c_orange_lbl.fill = fill_media
    c_orange_lbl.alignment = align_center
    c_orange_lbl.border = border_task
    sheet.cell(current_row, 4, "URGENTE").font = font_legend
    
    current_row += 1
    
    # Yellow (Cumplido)
    c_yellow_lbl = sheet.cell(current_row, 2, "AMARILLO :")
    c_yellow_lbl.font = font_legend
    c_yellow_lbl.fill = fill_completed
    c_yellow_lbl.alignment = align_center
    c_yellow_lbl.border = border_task
    sheet.cell(current_row, 4, "SE CUMPLIO LA TAREA").font = font_legend
    
    current_row += 1
    
    # Green (Baja)
    c_green_lbl = sheet.cell(current_row, 2, "VERDE :")
    c_green_lbl.font = font_legend
    c_green_lbl.fill = fill_baja
    c_green_lbl.alignment = align_center
    c_green_lbl.border = border_task
    sheet.cell(current_row, 4, "URGENTE - PERO PUEDE ESPERAR").font = font_legend
    
    # Set custom column widths for design aesthetics
    sheet.column_dimensions['A'].width = 3
    sheet.column_dimensions['B'].width = 18 # Names
    for col_let in ['C', 'D', 'E', 'F', 'G', 'H']:
        sheet.column_dimensions[col_let].width = 35 # Tasks
        
    # 4. Create No_Cumplidas sheet
    sheet_nc = wb.create_sheet(title="No_Cumplidas")
    sheet_nc.views.sheetView[0].showGridLines = True
    
    # Headers
    headers_nc = [
        "RESPONSABLE", 
        "ACTIVIDAD NO EJECUTADA", 
        "EN AVANCE O MOTIVO DE NO EJECUCIÓN", 
        "SOLUCIÓN Y/O MEDIDA ADOPTADA POR EL RESPONSABLE DE ÁREA"
    ]
    for col_idx, text in enumerate(headers_nc, start=2): # Column B (2), C (3), D (4), E (5)
        cell = sheet_nc.cell(row=3, column=col_idx, value=text)
        cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='medium', color='7F7F7F'),
            bottom=Side(style='medium', color='7F7F7F')
        )
        
    db_users = get_users()
    users_list = db_users if db_users else ["MARY CRUZ", "CPC.SHEYLA", "CPC.HECTOR"]
    
    pending_tasks_data = []
    
    for date_str in week_dates:
        for user_name in users_list:
            tasks = get_tasks(user_name, date_str)
            pending_tasks = [t for t in tasks if t['completed'] == 0]
            
            if not pending_tasks:
                continue
                
            reports = get_reports(user_name, date_str, date_str)
            report_suggestions = {}
            report_reasons = {}
            if reports:
                rep = reports[0]
                rep_unresolved = rep.get('unresolved_tasks', [])
                rep_alts = rep.get('alternatives_of_solution', [])
                for idx, ut in enumerate(rep_unresolved):
                    desc_key = ut.get('description', '').strip().upper()
                    if idx < len(rep_alts):
                        report_suggestions[desc_key] = rep_alts[idx]
                    report_reasons[desc_key] = ut.get('execution_status', '')
                        
            for task in pending_tasks:
                desc_key = task['description'].strip().upper()
                solution_val = report_suggestions.get(desc_key, "")
                reason_val = report_reasons.get(desc_key, "")
                
                # Extract solution text from dictionary or string representation
                if isinstance(solution_val, dict):
                    solution = solution_val.get("alternativa_solucion", "")
                elif isinstance(solution_val, str) and (solution_val.strip().startswith("{") or solution_val.strip().startswith("'{")):
                    import json
                    try:
                        parsed_val = json.loads(solution_val.replace("'", '"'))
                        if isinstance(parsed_val, dict):
                            solution = parsed_val.get("alternativa_solucion", "")
                        else:
                            solution = solution_val
                    except Exception:
                        solution = solution_val
                else:
                    solution = solution_val
                
                t_desc = task['description']
                t_time = task['time_info']
                task_str = t_desc
                if t_time and str(t_time).strip() != "" and str(t_time).upper() != "NONE":
                    task_str += f" - {t_time}"
                    
                pending_tasks_data.append({
                    "responsable": user_name,
                    "task": task_str,
                    "reason": reason_val,
                    "solution": solution
                })
                
    current_nc_row = 4
    if not pending_tasks_data:
        sheet_nc.merge_cells(start_row=current_nc_row, start_column=2, end_row=current_nc_row, end_column=5)
        cell = sheet_nc.cell(row=current_nc_row, column=2, value="¡Todas las tareas fueron completadas!")
        cell.font = Font(name="Calibri", size=11, bold=True, italic=True, color="1a7f37")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        border_side = Side(style='thin', color='D9D9D9')
        for col in range(2, 6):
            sheet_nc.cell(row=current_nc_row, column=col).border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    else:
        border_side = Side(style='thin', color='D9D9D9')
        border_task = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        font_task = Font(name="Calibri", size=10, color="000000")
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center")
        
        for item in pending_tasks_data:
            c_resp = sheet_nc.cell(row=current_nc_row, column=2, value=item["responsable"])
            c_resp.font = font_task
            c_resp.alignment = align_center
            c_resp.border = border_task
            
            c_task = sheet_nc.cell(row=current_nc_row, column=3, value=item["task"])
            c_task.font = font_task
            c_task.alignment = align_left
            c_task.border = border_task
            
            c_reason = sheet_nc.cell(row=current_nc_row, column=4, value=item["reason"])
            c_reason.font = font_task
            c_reason.alignment = align_left
            c_reason.border = border_task
            
            c_sol = sheet_nc.cell(row=current_nc_row, column=5, value=item["solution"])
            c_sol.font = font_task
            c_sol.alignment = align_left
            c_sol.border = border_task
            
            current_nc_row += 1
            
    sheet_nc.column_dimensions['A'].width = 3
    sheet_nc.column_dimensions['B'].width = 18 # Responsable
    sheet_nc.column_dimensions['C'].width = 45 # Actividad No Ejecutada
    sheet_nc.column_dimensions['D'].width = 45 # En Avance o Motivo de No Ejecución
    sheet_nc.column_dimensions['E'].width = 55 # Solución y/o Medida Adoptada
    
    # Save file
    wb.save(output_path)
